import pytest
import pypdf
import os
from path import RESOURCES_PATH
from openpyxl import load_workbook
import csv


def test_pdf():
    path = os.path.join(RESOURCES_PATH, "Python Testing with Pytest (Brian Okken) (1).pdf")
    reader = pypdf.PdfReader(path)
    text = reader.pages[1].extract_text()
    assert "Version: P1.0 (September 2017)" in text


def test_xlsx():
    workbook = load_workbook(os.path.join(RESOURCES_PATH, "file_example_XLSX_50.xlsx"))
    sheet = workbook.active
    assert sheet.cell(2, 3).value == "Abril"
    assert sheet.cell(17, 5).value == "France"


def test_csv():
    with open(os.path.join(RESOURCES_PATH, "file_example_csv.csv")) as f:
        rows = csv.reader(f)
        new_rows = [''.join(row).split(';') for row in rows]
        assert new_rows[2][1] == "Mara"
